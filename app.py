import streamlit as st
try:
    from google import genai as genai_new
    USE_NEW_SDK = True
except ImportError:
    import google.generativeai as genai
    USE_NEW_SDK = False
import pandas as pd
import json
import io
import re
import os
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
st.set_page_config(page_title="Logistics Billing Checker", page_icon="🚚", layout="wide")

api_key = os.getenv("GEMINI_API_KEY") or st.secrets.get("GEMINI_API_KEY", "")
if USE_NEW_SDK:
    _client = genai_new.Client(api_key=api_key)
    def _generate(prompt):
        return _client.models.generate_content(model="gemini-1.5-flash", contents=prompt).text
else:
    genai.configure(api_key=api_key)
    _old_model = genai.GenerativeModel("gemini-1.5-flash")
    def _generate(prompt):
        return _old__generate(prompt).text


# ═══════════════════════════════════════════════════════════════════════════════
# PINCODE DATABASE
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def load_pincode_db():
    if os.path.exists("pincode_db.csv"):
        df = pd.read_csv("pincode_db.csv", dtype=str, nrows=1)
        cols = [c.lower().strip() for c in df.columns]
        # Detect pincode column — handles Kaggle CSVs with different headers
        pincode_col = next((df.columns[i] for i, c in enumerate(cols)
                           if c in ["pincode", "pin", "pin_code", "postalcode",
                                    "postal_code", "pincodeno", "pin code"]), None)
        if pincode_col is None:
            return {}  # can not find pincode column — skip DB
        df = pd.read_csv("pincode_db.csv", dtype=str)
        df = df.rename(columns={pincode_col: "pincode"})
        # Detect state/district columns
        col_map = {}
        for orig_col in df.columns:
            lc = orig_col.lower().strip()
            if lc in ["statename", "state_name", "state", "statecode"]:
                col_map[orig_col] = "state"
            elif lc in ["districtname", "district_name", "district", "taluk"]:
                col_map[orig_col] = "district"
            elif lc in ["officename", "office_name", "officetype", "office"]:
                col_map[orig_col] = "office"
        df = df.rename(columns=col_map)
        # Keep only useful columns
        keep = [c for c in ["pincode","state","district","office"] if c in df.columns]
        df = df[keep].drop_duplicates("pincode")
        df["pincode"] = df["pincode"].str.strip().str.zfill(6)
        return df.set_index("pincode").to_dict("index")
    return {}

PINCODE_DB = load_pincode_db()

# City groupings — covers all major metros + tier-2 cities
PINCODE_TO_CITY = {}
CITY_PREFIXES = {
    "Mumbai":     ["400"],
    "Delhi":      ["110"],
    "Bangalore":  ["560"],
    "Chennai":    ["600"],
    "Kolkata":    ["700"],
    "Hyderabad":  ["500"],
    "Ahmedabad":  ["380"],
    "Pune":       ["411", "412"],
    "Jaipur":     ["302", "303"],
    "Lucknow":    ["226", "227"],
    "Ludhiana":   ["141"],
    "Chandigarh": ["160"],
    "Noida":      ["201"],
    "Gurugram":   ["122"],
    "Coimbatore": ["641", "642"],
    "Surat":      ["395"],
    "Kochi":      ["682", "683"],
    "Nagpur":     ["440"],
    "Indore":     ["452", "453"],
    "Bhopal":     ["462", "463"],
    "Patna":      ["800", "801"],
    "Bhubaneswar":["751", "752"],
    "Vishakhapatnam": ["530"],
    "Vadodara":   ["390"],
    "Agra":       ["282"],
    "Varanasi":   ["221"],
    "Rajkot":     ["360"],
    "Mysuru":     ["570"],
    "Thiruvananthapuram": ["695"],
    "Guwahati":   ["781"],
}

for city, prefixes in CITY_PREFIXES.items():
    for pfx in prefixes:
        for i in range(0, 100):
            PINCODE_TO_CITY[f"{pfx}{str(i).zfill(3)}"] = city

CITIES = list(CITY_PREFIXES.keys())

# ── ZONE MATRICES ─────────────────────────────────────────────────────────────
# Standard zone matrix (approximate Delhivery/Ecom/Shadowfax)
DELHIVERY_ZONES = {
    "Mumbai":    {"Mumbai":"A","Delhi":"B","Bangalore":"B","Chennai":"B","Kolkata":"C","Hyderabad":"B","Ahmedabad":"A","Pune":"A","Jaipur":"C","Lucknow":"C","Ludhiana":"C","Chandigarh":"C","Coimbatore":"C","Surat":"A","Noida":"B","Gurugram":"B","Kochi":"C","Nagpur":"B","Indore":"B","Bhopal":"C","Patna":"D","Bhubaneswar":"C","Vishakhapatnam":"C","Vadodara":"A","Agra":"C","Varanasi":"C","Rajkot":"B","Mysuru":"C","Thiruvananthapuram":"D","Guwahati":"E"},
    "Delhi":     {"Mumbai":"B","Delhi":"A","Bangalore":"C","Chennai":"C","Kolkata":"C","Hyderabad":"C","Ahmedabad":"B","Pune":"C","Jaipur":"A","Lucknow":"B","Ludhiana":"A","Chandigarh":"A","Coimbatore":"D","Surat":"C","Noida":"A","Gurugram":"A","Kochi":"D","Nagpur":"C","Indore":"B","Bhopal":"B","Patna":"B","Bhubaneswar":"C","Vishakhapatnam":"D","Vadodara":"B","Agra":"A","Varanasi":"B","Rajkot":"C","Mysuru":"D","Thiruvananthapuram":"D","Guwahati":"D"},
    "Bangalore": {"Mumbai":"B","Delhi":"C","Bangalore":"A","Chennai":"B","Kolkata":"C","Hyderabad":"B","Ahmedabad":"C","Pune":"B","Jaipur":"D","Lucknow":"D","Ludhiana":"D","Chandigarh":"D","Coimbatore":"B","Surat":"C","Noida":"C","Gurugram":"C","Kochi":"B","Nagpur":"C","Indore":"C","Bhopal":"D","Patna":"D","Bhubaneswar":"C","Vishakhapatnam":"C","Vadodara":"C","Agra":"D","Varanasi":"D","Rajkot":"D","Mysuru":"A","Thiruvananthapuram":"C","Guwahati":"E"},
    "Chennai":   {"Mumbai":"B","Delhi":"C","Bangalore":"B","Chennai":"A","Kolkata":"C","Hyderabad":"B","Ahmedabad":"C","Pune":"B","Jaipur":"D","Lucknow":"D","Ludhiana":"D","Chandigarh":"D","Coimbatore":"B","Surat":"C","Noida":"C","Gurugram":"C","Kochi":"B","Nagpur":"C","Indore":"D","Bhopal":"D","Patna":"D","Bhubaneswar":"C","Vishakhapatnam":"C","Vadodara":"C","Agra":"D","Varanasi":"D","Rajkot":"D","Mysuru":"B","Thiruvananthapuram":"B","Guwahati":"E"},
    "Kolkata":   {"Mumbai":"C","Delhi":"C","Bangalore":"C","Chennai":"C","Kolkata":"A","Hyderabad":"C","Ahmedabad":"D","Pune":"C","Jaipur":"D","Lucknow":"C","Ludhiana":"D","Chandigarh":"D","Coimbatore":"D","Surat":"D","Noida":"C","Gurugram":"C","Kochi":"D","Nagpur":"C","Indore":"D","Bhopal":"D","Patna":"B","Bhubaneswar":"B","Vishakhapatnam":"C","Vadodara":"D","Agra":"C","Varanasi":"B","Rajkot":"E","Mysuru":"D","Thiruvananthapuram":"D","Guwahati":"C"},
    "Hyderabad": {"Mumbai":"B","Delhi":"C","Bangalore":"B","Chennai":"B","Kolkata":"C","Hyderabad":"A","Ahmedabad":"C","Pune":"B","Jaipur":"D","Lucknow":"D","Ludhiana":"D","Chandigarh":"D","Coimbatore":"C","Surat":"C","Noida":"C","Gurugram":"C","Kochi":"C","Nagpur":"B","Indore":"C","Bhopal":"C","Patna":"D","Bhubaneswar":"C","Vishakhapatnam":"B","Vadodara":"C","Agra":"D","Varanasi":"D","Rajkot":"D","Mysuru":"C","Thiruvananthapuram":"C","Guwahati":"E"},
    "Ahmedabad": {"Mumbai":"A","Delhi":"B","Bangalore":"C","Chennai":"C","Kolkata":"D","Hyderabad":"C","Ahmedabad":"A","Pune":"B","Jaipur":"B","Lucknow":"C","Ludhiana":"C","Chandigarh":"C","Coimbatore":"D","Surat":"A","Noida":"B","Gurugram":"B","Kochi":"D","Nagpur":"C","Indore":"B","Bhopal":"B","Patna":"D","Bhubaneswar":"D","Vishakhapatnam":"D","Vadodara":"A","Agra":"C","Varanasi":"D","Rajkot":"A","Mysuru":"D","Thiruvananthapuram":"E","Guwahati":"E"},
    "Pune":      {"Mumbai":"A","Delhi":"C","Bangalore":"B","Chennai":"B","Kolkata":"C","Hyderabad":"B","Ahmedabad":"B","Pune":"A","Jaipur":"C","Lucknow":"D","Ludhiana":"D","Chandigarh":"D","Coimbatore":"C","Surat":"B","Noida":"C","Gurugram":"C","Kochi":"C","Nagpur":"B","Indore":"C","Bhopal":"C","Patna":"D","Bhubaneswar":"C","Vishakhapatnam":"C","Vadodara":"B","Agra":"D","Varanasi":"D","Rajkot":"C","Mysuru":"C","Thiruvananthapuram":"D","Guwahati":"E"},
}
# Fill remaining cities with reasonable defaults
for city in CITIES:
    if city not in DELHIVERY_ZONES:
        row = {}
        for other in CITIES:
            if other == city: row[other] = "A"
            elif other in ["Mumbai","Delhi","Bangalore","Chennai","Kolkata","Hyderabad"]: row[other] = "C"
            else: row[other] = "D"
        DELHIVERY_ZONES[city] = row

# BlueDart has slightly different zone definitions for some corridors
BLUEDART_ZONES = {k: dict(v) for k, v in DELHIVERY_ZONES.items()}
# BlueDart specific overrides (example)
if "Mumbai" in BLUEDART_ZONES:
    BLUEDART_ZONES["Mumbai"]["Delhi"] = "B"
    BLUEDART_ZONES["Mumbai"]["Ahmedabad"] = "A"
if "Delhi" in BLUEDART_ZONES:
    BLUEDART_ZONES["Delhi"]["Mumbai"] = "B"
    BLUEDART_ZONES["Delhi"]["Noida"] = "A"

PROVIDER_MATRICES = {
    "delhivery": DELHIVERY_ZONES,
    "bluedart": BLUEDART_ZONES,
    "blue dart": BLUEDART_ZONES,
    "ecom express": DELHIVERY_ZONES,
    "ecom": DELHIVERY_ZONES,
    "shadowfax": DELHIVERY_ZONES,
}

DISTRICT_TO_CITY = {
    "Mumbai": "Mumbai", "Mumbai City": "Mumbai", "Mumbai Suburban": "Mumbai",
    "New Delhi": "Delhi", "North Delhi": "Delhi", "South Delhi": "Delhi",
    "East Delhi": "Delhi", "West Delhi": "Delhi", "Central Delhi": "Delhi",
    "Bangalore": "Bangalore", "Bengaluru": "Bangalore", "Bengaluru Urban": "Bangalore",
    "Chennai": "Chennai", "Kolkata": "Kolkata", "Hyderabad": "Hyderabad",
    "Ahmedabad": "Ahmedabad", "Pune": "Pune", "Jaipur": "Jaipur",
    "Lucknow": "Lucknow", "Ludhiana": "Ludhiana", "Chandigarh": "Chandigarh",
    "Gautam Buddha Nagar": "Noida", "Gurgaon": "Gurugram", "Gurugram": "Gurugram",
    "Coimbatore": "Coimbatore", "Surat": "Surat", "Ernakulam": "Kochi",
    "Nagpur": "Nagpur", "Indore": "Indore", "Bhopal": "Bhopal",
    "Patna": "Patna", "Khordha": "Bhubaneswar", "Visakhapatnam": "Vishakhapatnam",
    "Vadodara": "Vadodara", "Agra": "Agra", "Varanasi": "Varanasi",
    "Rajkot": "Rajkot", "Mysore": "Mysuru", "Mysuru": "Mysuru",
    "Thiruvananthapuram": "Thiruvananthapuram", "Kamrup Metropolitan": "Guwahati",
}

def pin_to_city(pin):
    pin = str(pin).strip().zfill(6)
    # Direct match
    if pin in PINCODE_TO_CITY:
        return PINCODE_TO_CITY[pin]
    # 4-digit prefix match
    if pin[:4] + "0" * 2 in PINCODE_TO_CITY:
        return PINCODE_TO_CITY[pin[:4] + "0" * 2]
    # 3-digit prefix match (broad)
    key = pin[:3] + "0" * 3
    if key in PINCODE_TO_CITY:
        return PINCODE_TO_CITY[key]
    # Pincode DB fallback
    if pin in PINCODE_DB:
        dist = PINCODE_DB[pin].get("district", "")
        return DISTRICT_TO_CITY.get(dist)
    return None

def get_correct_zone(origin_pin, dest_pin, provider="delhivery"):
    oc = pin_to_city(origin_pin)
    dc = pin_to_city(dest_pin)
    if not oc or not dc:
        return None
    matrix = PROVIDER_MATRICES.get(provider.lower(), DELHIVERY_ZONES)
    return matrix.get(oc, {}).get(dc)


# ═══════════════════════════════════════════════════════════════════════════════
# WEIGHT & RATE CALCULATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_billable_weight(actual_kg, L=None, W=None, H=None, dim_divisor=None, slab="ceil_500g"):
    """Returns (billable_weight, method_string)"""
    vol = round((L * W * H) / dim_divisor, 3) if (dim_divisor and L and W and H) else None
    base = max(actual_kg, vol) if vol else actual_kg
    method = "dimensional" if (vol and vol > actual_kg) else "actual"
    if slab == "ceil_500g":
        return math.ceil(base * 2) / 2, method
    elif slab == "ceil_100g":
        return math.ceil(base * 10) / 10, method
    return round(base, 3), method


def normalize_contract(c):
    """Normalize contract zone keys regardless of what the AI named them."""
    out = dict(c)
    out["zones"] = {}
    for zone, rate_obj in c.get("zones", {}).items():
        kv = {k.lower().replace(" ", "_").replace("-", "_"): v
              for k, v in rate_obj.items() if isinstance(v, (int, float))}
        u500 = u1kg = pkg = None
        for k, v in kv.items():
            if any(x in k for x in ["upto_500", "below_500", "first_500", "0_500", "half_kg", "0_5kg"]):
                u500 = v
            elif any(x in k for x in ["1kg", "1_kg", "500_to_1", "half_to_1", "500g_1"]):
                u1kg = v
            elif any(x in k for x in ["per_kg", "each_kg", "above_1", "beyond_1", "additional"]):
                pkg = v
        sv = sorted(kv.values())
        if u500 is None and sv:          u500 = sv[0]
        if u1kg is None and len(sv) > 1: u1kg = sv[1]
        if pkg  is None and len(sv) > 2: pkg  = sv[2]
        out["zones"][zone] = {
            "upto_500g": u500 or 0,
            "500g_to_1kg": u1kg or 0,
            "per_kg_above_1kg": pkg or 0
        }
    return out


def fwd_charge(zone, wt, contract):
    r = contract.get("zones", {}).get(zone, {})
    if wt <= 0.5:   return r.get("upto_500g", 0)
    elif wt <= 1.0: return r.get("500g_to_1kg", 0)
    else:           return r.get("500g_to_1kg", 0) + (wt - 1.0) * r.get("per_kg_above_1kg", 0)


ALL_SURCHARGE_FIELDS = [
    "fuel_surcharge", "oda_charge", "docket_charge", "handling_charge",
    "peak_surcharge", "special_handling", "remote_area_charge",
    "volumetric_surcharge", "sunday_surcharge", "holiday_surcharge",
    "re_attempt_charge", "address_correction_charge", "other_surcharges",
]


# ═══════════════════════════════════════════════════════════════════════════════
# CHECKING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def check_invoice(items, contract, manifest_df=None):
    contract = normalize_contract(contract)
    provider = contract.get("provider", "delhivery").lower()
    slab     = contract.get("weight_slab_rounding", "ceil_500g")
    div      = contract.get("dimensional_weight_divisor")
    tol      = contract.get("weight_tolerance_kg", 0.05)
    contracted = [s.lower().strip().replace(" ", "_")
                  for s in contract.get("contracted_surcharges", [])]

    disc = []
    seen_awb = {}

    # Build manifest lookup
    mlu = {}
    if manifest_df is not None:
        awb_col = next((c for c in manifest_df.columns if "awb" in c.lower()), None)
        if awb_col:
            for _, row in manifest_df.iterrows():
                mlu[str(row[awb_col]).strip()] = row.to_dict()

    for idx, item in enumerate(items):
        awb  = str(item.get("awb_number", f"ROW_{idx}")).strip()
        zone = str(item.get("zone") or "").upper().strip()
        bwt  = float(item.get("billed_weight_kg") or 0)
        orig = str(item.get("origin_pincode") or "").strip()
        dest = str(item.get("dest_pincode") or "").strip()

        # ── 1. Duplicate AWB ──────────────────────────────────────────────────
        if awb in seen_awb:
            total = float(item.get("total") or 0)
            disc.append({
                "AWB": awb, "Error Type": "Duplicate AWB",
                "Description": f"AWB {awb} billed again (first at row {seen_awb[awb]+1})",
                "Billed (₹)": total, "Correct (₹)": 0, "Overcharge (₹)": total
            })
        else:
            seen_awb[awb] = idx

        # ── 2. Weight vs Manifest (actual weight + dimensional) ───────────────
        if awb in mlu:
            mr = mlu[awb]
            aw = float(mr.get("Actual_Weight_Kg") or mr.get("actual_weight_kg") or 0)
            L  = float(mr.get("Length_cm") or mr.get("length_cm") or 0) or None
            W  = float(mr.get("Width_cm")  or mr.get("width_cm")  or 0) or None
            H  = float(mr.get("Height_cm") or mr.get("height_cm") or 0) or None
            if aw > 0:
                correct_wt, method = get_billable_weight(aw, L, W, H, div, slab)
                if bwt > correct_wt + tol:
                    disc.append({
                        "AWB": awb, "Error Type": "Weight Overcharge",
                        "Description": f"Billed {bwt}kg, correct {correct_wt}kg ({method}). Actual={aw}kg",
                        "Billed (₹)": bwt, "Correct (₹)": correct_wt,
                        "Overcharge (₹)": round(bwt - correct_wt, 3)
                    })
                bwt = correct_wt  # use corrected weight for downstream checks

        # ── 3. Zone Mismatch ─────────────────────────────────────────────────
        corrected_zone = zone
        if orig and dest and zone:
            cz = get_correct_zone(orig, dest, provider)
            if cz and cz != zone:
                bf = float(item.get("forward_charge") or 0)
                cf = round(fwd_charge(cz, bwt, contract), 2)
                disc.append({
                    "AWB": awb, "Error Type": "Zone Mismatch",
                    "Description": f"Billed Zone {zone}, correct Zone {cz} ({orig}→{dest})",
                    "Billed (₹)": bf, "Correct (₹)": cf,
                    "Overcharge (₹)": round(bf - cf, 2)
                })
                corrected_zone = cz

        # ── 4. Rate Deviation ─────────────────────────────────────────────────
        bf = float(item.get("forward_charge") or 0)
        if bf > 0 and corrected_zone in contract.get("zones", {}):
            cf = round(fwd_charge(corrected_zone, bwt, contract), 2)
            if cf > 0 and abs(bf - cf) / cf > 0.05:
                disc.append({
                    "AWB": awb, "Error Type": "Rate Deviation",
                    "Description": f"Charged ₹{bf}, contracted ₹{cf} (Zone {corrected_zone}, {bwt}kg)",
                    "Billed (₹)": bf, "Correct (₹)": cf,
                    "Overcharge (₹)": round(bf - cf, 2)
                })

        # ── 5. RTO Overcharge ─────────────────────────────────────────────────
        rto = float(item.get("rto_charge") or 0)
        if rto > 0 and corrected_zone in contract.get("rto_rates", {}):
            crto = contract["rto_rates"][corrected_zone]
            if rto > crto * 1.05:
                disc.append({
                    "AWB": awb, "Error Type": "RTO Overcharge",
                    "Description": f"RTO ₹{rto}, contracted ₹{crto} for Zone {corrected_zone}",
                    "Billed (₹)": rto, "Correct (₹)": crto,
                    "Overcharge (₹)": round(rto - crto, 2)
                })

        # ── 6. COD Fee ────────────────────────────────────────────────────────
        cod_amt = float(item.get("cod_amount") or 0)
        cod_fee = float(item.get("cod_fee") or 0)
        if cod_fee > 0:
            if cod_amt == 0:
                disc.append({
                    "AWB": awb, "Error Type": "COD Fee on Prepaid",
                    "Description": f"COD fee ₹{cod_fee} charged but COD amount = ₹0 (prepaid order)",
                    "Billed (₹)": cod_fee, "Correct (₹)": 0, "Overcharge (₹)": cod_fee
                })
            else:
                correct_cod = max(
                    contract.get("cod_fee_minimum", 25),
                    cod_amt * contract.get("cod_fee_percent", 1.5) / 100
                )
                if cod_fee > correct_cod * 1.05:
                    disc.append({
                        "AWB": awb, "Error Type": "COD Fee Overcharge",
                        "Description": f"COD fee ₹{cod_fee}, contracted ₹{correct_cod:.2f} ({contract.get('cod_fee_percent')}% of ₹{cod_amt})",
                        "Billed (₹)": cod_fee, "Correct (₹)": round(correct_cod, 2),
                        "Overcharge (₹)": round(cod_fee - correct_cod, 2)
                    })

        # ── 7. Non-Contracted Surcharges ──────────────────────────────────────
        for field in ALL_SURCHARGE_FIELDS:
            amt = float(item.get(field) or 0)
            if amt > 0:
                fn = field.lower().replace(" ", "_")
                if not any(c in fn or fn in c for c in contracted):
                    disc.append({
                        "AWB": awb, "Error Type": "Non-Contracted Surcharge",
                        "Description": f"'{field}' ₹{amt} not in contracted surcharges list",
                        "Billed (₹)": amt, "Correct (₹)": 0, "Overcharge (₹)": amt
                    })

    tb = sum(float(i.get("total") or 0) for i in items)
    to = sum(max(0, d["Overcharge (₹)"]) for d in disc)
    ebt = {}
    for d in disc:
        ebt[d["Error Type"]] = round(ebt.get(d["Error Type"], 0) + max(0, d["Overcharge (₹)"]), 2)

    return disc, {
        "total_items": len(items),
        "total_billed": round(tb, 2),
        "total_overcharge": round(to, 2),
        "verified_amount": round(tb - to, 2),
        "error_count": len(disc),
        "errors_by_type": ebt,
        "provider": contract.get("provider", "Unknown"),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# AI EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

INV_PROMPT = """Extract ALL shipment rows from this logistics invoice.
Return a JSON array ONLY — no markdown, no explanation, no preamble.

Each object must have exactly:
{"awb_number":"string","date":"YYYY-MM-DD","origin_pincode":"6 digits",
"dest_pincode":"6 digits","billed_weight_kg":number,"zone":"A-E or null",
"forward_charge":number,"rto_charge":number,"cod_amount":number,"cod_fee":number,
"fuel_surcharge":number,"oda_charge":number,"docket_charge":number,
"handling_charge":number,"peak_surcharge":number,"special_handling":number,
"remote_area_charge":number,"other_surcharges":number,"total":number}

Rules: use 0 for missing numbers. "500g"→0.5. Start with [ end with ]."""

CON_PROMPT = """Extract the rate card from this logistics contract.
Return JSON only — no markdown, no explanation.

{"provider":"Delhivery/BlueDart/Ecom Express/Shadowfax",
"zones":{"A":{"upto_500g":n,"500g_to_1kg":n,"per_kg_above_1kg":n},"B":{...},"C":{...},"D":{...},"E":{...}},
"rto_rates":{"A":n,"B":n,"C":n,"D":n,"E":n},
"cod_fee_percent":n,"cod_fee_minimum":n,
"contracted_surcharges":["exact names of allowed surcharges"],
"weight_slab_rounding":"ceil_500g",
"weight_tolerance_kg":n,
"dimensional_weight_divisor":n_or_null}"""


def safe_json(text):
    text = re.sub(r"```json|```", "", text.strip()).strip()
    try:
        return json.loads(text)
    except:
        candidates = []
        if "[" in text: candidates.append(text.find("["))
        if "{" in text: candidates.append(text.find("{"))
        if candidates:
            s = min(candidates)
            e = max(text.rfind("]"), text.rfind("}"))
            if 0 <= s < e:
                try:
                    return json.loads(text[s:e+1])
                except:
                    pass
    fix = _generate(f"Fix this broken JSON, return only valid JSON:\n{text[:3000]}")
    return json.loads(re.sub(r"```json|```", "", fix.strip()).strip())


def ai_extract_invoice(content):
    all_items = []
    for chunk in [content[i:i+50000] for i in range(0, len(content), 50000)]:
        r = _generate(f"{INV_PROMPT}\n\nDATA:\n{chunk}")
        parsed = safe_json(r)
        if isinstance(parsed, list):
            all_items.extend(parsed)
    return all_items


def ai_extract_contract(content):
    r = _generate(f"{CON_PROMPT}\n\nCONTRACT:\n{content}")
    return safe_json(r)


# ═══════════════════════════════════════════════════════════════════════════════
# FILE READING
# ═══════════════════════════════════════════════════════════════════════════════

def read_file(f):
    n = f.name.lower()
    if n.endswith(".json"):
        return f.read().decode("utf-8")
    elif n.endswith(".csv"):
        return pd.read_csv(f).to_csv(index=False)
    elif n.endswith((".xlsx", ".xls")):
        return pd.read_excel(f).to_csv(index=False)
    elif n.endswith(".pdf"):
        import pdfplumber
        text = ""
        with pdfplumber.open(f) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: text += t + "\n"
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        text += " | ".join(str(c or "") for c in row) + "\n"
        return text
    return f.read().decode("utf-8", errors="ignore")


def read_manifest(f):
    return pd.read_csv(f) if f.name.lower().endswith(".csv") else pd.read_excel(f)


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def generate_excel(items, disc, summary):
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Billing Check — {summary.get('provider','')} — Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    for r, (lbl, val, col) in enumerate([
        ("Total Items Checked",   summary["total_items"],    None),
        ("Total Billed (₹)",      summary["total_billed"],   None),
        ("Verified Payable (₹)",  summary["verified_amount"],"C6EFCE"),
        ("Overcharges Found (₹)", summary["total_overcharge"],"FFC7CE"),
        ("Savings %", f"{summary['total_overcharge']/max(summary['total_billed'],1)*100:.1f}%", "FFC7CE"),
        ("Total Errors Found",    summary["error_count"],    None),
    ], start=3):
        ws.cell(r, 1, lbl).font = Font(bold=True)
        c = ws.cell(r, 2, val)
        if col: c.fill = PatternFill("solid", fgColor=col)

    ws.cell(11, 1, "Breakdown by Error Type").font = Font(bold=True)
    for r, (t, a) in enumerate(summary["errors_by_type"].items(), 12):
        ws.cell(r, 1, t)
        ws.cell(r, 2, f"₹{a:,.2f}").fill = PatternFill("solid", fgColor="FFC7CE")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    # Sheet 2: Verified Payout
    ws2 = wb.create_sheet("Verified Payout")
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    hdrs = ["AWB", "Date", "Origin", "Dest", "Weight(kg)", "Zone",
            "Total Billed(₹)", "Overcharge(₹)", "Verified(₹)", "Status", "Issues"]
    ws2.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        ws2.cell(1, c).fill = hf
        ws2.cell(1, c).font = hfont
        ws2.column_dimensions[get_column_letter(c)].width = 18

    dm = {}
    for d in disc:
        dm.setdefault(d["AWB"], []).append(d)

    for item in items:
        awb  = item.get("awb_number", "")
        errs = dm.get(awb, [])
        oc   = sum(max(0, e["Overcharge (₹)"]) for e in errs)
        tot  = float(item.get("total") or 0)
        ws2.append([
            awb, item.get("date",""), item.get("origin_pincode",""),
            item.get("dest_pincode",""), item.get("billed_weight_kg", 0),
            item.get("zone",""), round(tot, 2), round(oc, 2),
            round(tot - oc, 2),
            "⚠️ DISPUTED" if errs else "✅ CLEAN",
            "; ".join(set(e["Error Type"] for e in errs))
        ])
        if errs:
            for c in range(1, 12):
                ws2.cell(ws2.max_row, c).fill = PatternFill("solid", fgColor="FFC7CE")

    # Sheet 3: Discrepancy Report
    ws3 = wb.create_sheet("Discrepancy Report")
    dh = ["AWB", "Error Type", "Description", "Billed(₹)", "Correct(₹)", "Overcharge(₹)"]
    ws3.append(dh)
    for c in range(1, 7):
        ws3.cell(1, c).fill = PatternFill("solid", fgColor="C00000")
        ws3.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws3.column_dimensions["C"].width = 65
    for col in ["A", "B", "D", "E", "F"]:
        ws3.column_dimensions[col].width = 20

    for d in sorted(disc, key=lambda x: -max(0, x["Overcharge (₹)"])):
        ws3.append([
            d["AWB"], d["Error Type"], d["Description"],
            d["Billed (₹)"], d["Correct (₹)"], d["Overcharge (₹)"]
        ])
        ws3.cell(ws3.max_row, 6).fill = PatternFill("solid", fgColor="FFC7CE")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# SAMPLE CONTRACT (for sidebar download)
# ═══════════════════════════════════════════════════════════════════════════════

SAMPLE_CONTRACT = {
    "provider": "Delhivery",
    "zones": {
        "A": {"upto_500g": 38, "500g_to_1kg": 42, "per_kg_above_1kg": 20},
        "B": {"upto_500g": 42, "500g_to_1kg": 47, "per_kg_above_1kg": 22},
        "C": {"upto_500g": 48, "500g_to_1kg": 54, "per_kg_above_1kg": 26},
        "D": {"upto_500g": 55, "500g_to_1kg": 62, "per_kg_above_1kg": 30},
        "E": {"upto_500g": 65, "500g_to_1kg": 74, "per_kg_above_1kg": 35}
    },
    "rto_rates": {"A": 30, "B": 35, "C": 40, "D": 45, "E": 55},
    "cod_fee_percent": 1.5, "cod_fee_minimum": 25,
    "contracted_surcharges": ["fuel_surcharge", "docket_charge", "oda_charge"],
    "weight_slab_rounding": "ceil_500g",
    "weight_tolerance_kg": 0.05,
    "dimensional_weight_divisor": None
}


# ═══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("🚚 Billing Checker")
    st.markdown("---")
    st.markdown("**Checks performed:**")
    for c in [
        "✅ Duplicate AWB",
        "✅ Zone Mismatch",
        "✅ Rate Deviation",
        "✅ RTO Overcharge",
        "✅ COD Fee Overcharge",
        "✅ COD Fee on Prepaid",
        "✅ Weight Overcharge (with manifest)",
        "✅ Dimensional Weight (BlueDart etc.)",
        "✅ Non-Contracted Surcharges",
    ]:
        st.markdown(c)
    st.markdown("---")
    st.markdown("**Providers:** Delhivery · BlueDart · Ecom Express · Shadowfax")
    st.markdown("---")
    st.download_button(
        "⬇️ Sample Contract (Delhivery)",
        json.dumps(SAMPLE_CONTRACT, indent=2),
        "delhivery_contract.json",
        "application/json"
    )
    st.markdown("---")
    if os.path.exists("pincode_db.csv"):
        db = load_pincode_db()
        st.success(f"📍 Pincode DB: {len(db):,} pincodes loaded")
    else:
        st.warning("⚠️ pincode_db.csv not found\nRun: python download_pincodes.py")


st.title("🚚 Logistics Billing Checker")
st.markdown("##### AI-powered invoice verification · Minutes, not days")
st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    st.markdown("### 📄 Invoice")
    inv_file = st.file_uploader("CSV, Excel, or PDF", type=["csv", "xlsx", "xls", "pdf"], key="inv")
    if inv_file: st.success(f"✅ {inv_file.name}")

with col2:
    st.markdown("### 📋 Contract / Rate Card")
    con_file = st.file_uploader("JSON, CSV, Excel, or PDF",
                                 type=["json", "csv", "xlsx", "xls", "pdf"], key="con")
    if con_file: st.success(f"✅ {con_file.name}")

with st.expander("➕ Optional: Shipment Manifest (enables weight + dimensional checks)"):
    st.caption(
        "Upload your WMS/OMS export to enable weight verification.\n\n"
        "Required columns: **AWB_Number, Actual_Weight_Kg**\n"
        "Optional columns: **Length_cm, Width_cm, Height_cm** (for dimensional weight)\n\n"
        "All other checks (zone, rate, COD, RTO, surcharges) work without this file."
    )
    man_file = st.file_uploader("Manifest CSV or Excel", type=["csv", "xlsx", "xls"], key="man")
    if man_file:
        st.success(f"✅ {man_file.name} — weight + dimensional checks enabled")
    else:
        st.info("No manifest uploaded — weight checks skipped, all other checks active")

st.markdown("---")

if st.button("🚀 Run Billing Check", type="primary", disabled=not (inv_file and con_file)):
    with st.status("Processing...", expanded=True) as status:
        st.write("📂 Reading uploaded files...")
        inv_text = read_file(inv_file)
        con_text = read_file(con_file)
        man_df   = read_manifest(man_file) if man_file else None

        st.write("📋 Extracting contract rates...")
        con_data = (json.loads(con_text)
                    if con_file.name.endswith(".json")
                    else ai_extract_contract(con_text))

        st.write("📄 Extracting invoice line items with AI...")
        inv_items = ai_extract_invoice(inv_text)
        st.write(f"   → **{len(inv_items)} line items** extracted")
        if man_df is not None:
            st.write(f"   → Manifest loaded: **{len(man_df)} rows**")

        st.write("🔍 Running all checks against contracted rates...")
        disc, summary = check_invoice(inv_items, con_data, man_df)
        st.write(f"   → **{summary['error_count']} discrepancies** found")

        st.write("📊 Building payout Excel file...")
        payout = generate_excel(inv_items, disc, summary)
        status.update(label="✅ Analysis complete!", state="complete")

    st.markdown("---")
    st.markdown(f"## 📊 Results — {summary.get('provider', '')}")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Items Checked", f"{summary['total_items']:,}")
    m2.metric("Total Billed", f"₹{summary['total_billed']:,.0f}")
    m3.metric("Overcharges Found", f"₹{summary['total_overcharge']:,.0f}",
              delta=f"-{summary['total_overcharge']/max(summary['total_billed'],1)*100:.1f}%",
              delta_color="inverse")
    m4.metric("Errors Found", summary["error_count"])

    ca, cb = st.columns(2)
    with ca:
        st.markdown("#### By Error Type")
        if summary["errors_by_type"]:
            st.bar_chart(
                pd.DataFrame(list(summary["errors_by_type"].items()),
                             columns=["Type", "₹"]).set_index("Type")
            )
        else:
            st.success("🎉 No overcharges found!")
    with cb:
        st.markdown("#### Summary")
        st.dataframe(pd.DataFrame([
            {"": "Total Billed",      "Amount (₹)": summary["total_billed"]},
            {"": "Verified Payable",  "Amount (₹)": summary["verified_amount"]},
            {"": "Overcharges",       "Amount (₹)": summary["total_overcharge"]},
        ]), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("#### 🚨 Discrepancy Report")
    if disc:
        df = pd.DataFrame(disc).sort_values("Overcharge (₹)", ascending=False)
        st.dataframe(df, use_container_width=True, hide_index=True,
                     column_config={c: st.column_config.NumberColumn(format="₹%.2f")
                                    for c in ["Billed (₹)", "Correct (₹)", "Overcharge (₹)"]})
    else:
        st.success("✅ Invoice matches contracted rates — no discrepancies found!")

    st.markdown("---")
    st.download_button(
        "⬇️ Download Payout File (Excel)",
        payout,
        f"payout_{summary.get('provider','').replace(' ','_')}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
    st.caption("3 sheets: Summary · Verified Payout (disputes highlighted) · Discrepancy Report")

elif not (inv_file and con_file):
    st.info("Upload invoice + contract above to begin. Download a sample contract from the sidebar →")